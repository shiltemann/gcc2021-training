import openpyxl
import argparse
import os
import re
import json
import pytz
import datetime
import sys

parser = argparse.ArgumentParser(description='Process a spreadsheet')
parser.add_argument('spreadsheet', type=argparse.FileType('r'),
                    help='A spreadsheet to process (in our special format!)')
parser.add_argument('--dry-run', action='store_true')

args = parser.parse_args()

wb = openpyxl.load_workbook(args.spreadsheet.name)
ws = wb["Messages"]

channelMap = {
    "#announcements": "announce",
    "#bofs": "bofs",
    "#cofest": "cofest",
    "#remo-assistants": "remo",
    "#social": "social",
}

for row in range(2, 700):
    if (
        ws.cell(row=row, column=5).value == "twitter"
        or ws.cell(row=row, column=5).value is None
        or ws.cell(row=row, column=1).value
    ):
        continue

    data = []
    for col in range(2, 7):
        data.append(ws.cell(row=row, column=col).value)

    (hours, minutes) = map(int, data[1].split(":"))
    dt = data[0]
    fixed = datetime.datetime(dt.year, dt.month, dt.day, hours, minutes, 0)
    loc_dt = fixed.astimezone(pytz.timezone(data[2]))
    fn = loc_dt.strftime("%Y-%m-%dT%H:%M:%S%z")

    channel = data[3]
    message = data[4]
    preview = re.sub("[^A-Za-z0-9_ -]", "", message.replace(" ", "-"))
    preview = re.sub("\[([^]]*)\]\(([^)]*)\)", r"\1", preview)[0:15]
    fullfn = os.path.join(
        "scheduled", "gcc", channelMap[channel], f"{fn}-{preview}.json"
    )

    text = re.sub("<", "&lt;", message)
    text = re.sub(">", "&gt;", text)
    text = re.sub("&", "&amp;", text)
    text = re.sub("(@U[0-9A-Z]*)", r"<\1>", text)
    text = re.sub("\[([^]]*)\]\(([^)]*)\)", r"<\2|\1>", text)

    slack_message = {
        "blocks": [
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": text,
                },
            }
        ]
    }

    if channel != "#remo-assistants":
        slack_message["blocks"].append(
            {
                "type": "actions",
                "elements": [
                    {
                        "type": "button",
                        "text": {
                            "type": "plain_text",
                            "text": "Schedule :calendar:",
                            "emoji": True,
                        },
                        "url": "https://gcc2021.sched.com/",
                    },
                    {
                        "type": "button",
                        "text": {
                            "type": "plain_text",
                            "text": "Code of Conduct",
                            "emoji": True,
                        },
                        "url": "https://www.vibconferences.be/code-of-conduct",
                    },
                    {
                        "type": "button",
                        "text": {
                            "type": "plain_text",
                            "text": "Remo|Main",
                            "emoji": True,
                        },
                        "url": "https://live.remo.co/e/galaxy-main-event",
                    },
                    {
                        "type": "button",
                        "text": {
                            "type": "plain_text",
                            "text": "Remo|Posters",
                            "emoji": True,
                        },
                        "url": "https://live.remo.co/e/galaxy-poster-session",
                    },
                    {
                        "type": "button",
                        "text": {
                            "type": "plain_text",
                            "text": "Remo|Social",
                            "emoji": True,
                        },
                        "url": "https://live.remo.co/e/galaxy-social-event",
                    },
                ],
            }
        )

        slack_message["blocks"].append(
            {
                "type": "context",
                "elements": [
                    {
                        "type": "plain_text",
                        "text": "Author: GCC Organisers",
                    }
                ],
            }
        )

    if not args.dry_run:
        with open(fullfn, 'w') as handle:
            json.dump(slack_message, handle, indent=2)
    else:
        print(fullfn)
        print(json.dumps(slack_message))
